import win32com.client
import utils
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Powerpoint:
    def __init__(self, template_path, output_path, template_index=1):
        self.template_index = template_index
        self.template_path = template_path
        self.output_path = output_path
        self.instance = win32com.client.Dispatch("PowerPoint.Application")
        self.presentation = self.instance.Presentations.Open(template_path, WithWindow=False)

    def close(self):
        # Remove template slide
        self.presentation.Slides(self.template_index).Delete()
        # Save the presentation
        self.presentation.SaveAs(self.output_path)
        # Close the presentation
        self.presentation.Close()
        # Quit the PowerPoint application
        self.instance.Quit()
        print("Quit out cleanly")

    def new_slide(self):
        return self.presentation.Slides(self.template_index).Duplicate()

    def update_text(self, slide, element_name, new_text):
        element = slide.Shapes(element_name)
        if not element.HasTextFrame:
            raise Exception(f"Tried to set text for element '{element}' which does not have a text field")
        element.TextFrame.TextRange.Text = new_text

    def update_other(self, slide, element_name, data, function_name):
        try:
            function = getattr(utils, function_name)
        except:
            raise Exception(f"Failed to get function '{function_name}'... are you sure it is in utils.py?")
        element = slide.Shapes(element_name)
        function(slide, element, data)

    def pivot_input_data(self, data):
        transposed_list = []
        # Get the length of the sublists
        sublist_length = len(data[0])
        # Iterate over the sublists, transposing the elements
        for i in range(sublist_length):
            transposed_sublist = []
            # Iterate over the main list and extract elements at index i
            for sublist in data:
                transposed_sublist.append(sublist[i])
            transposed_list.append(transposed_sublist)
        return transposed_list

    def set_chart_data(self, slide, chart_name : str, data):
        values_for_chart = self.pivot_input_data(data)\
        # Identify the chart shape on the slide
        shape = slide.Shapes(chart_name)
        # Retrieve the chart object from the shape
        chart = shape.Chart
        try:
            # Load workbook
            workbook_path = shape.LinkFormat.SourceFullName
            print(workbook_path)
            wb = load_workbook(workbook_path)
            ws = wb.active
            table = ws.tables['data']
            # Clear existing data
            ws.delete_rows(2,1000) # Assumption is no chart will utilize more than 1000 rows, can increase this if necessary
            # Resize the table
            num_rows = len(values_for_chart)
            num_cols = len(values_for_chart[0])
            range_string = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
            table.ref = range_string        
            # Update the values in the range
            for row_index, row in enumerate(ws[range_string]):
                # Skip header row
                if row_index == 0:
                    continue
                for col_index, cell in enumerate(row):
                    cell.value = values_for_chart[row_index][col_index]

            # Close the workbook
            chart.ChartData.Workbook.Close()
        except Exception as e:
            print("An error occurred:", str(e))
